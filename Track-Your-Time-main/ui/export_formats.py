"""
Multi-Format Export for Time Tracker
Supports: CSV, JSON, Excel, PDF
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import json
import csv
from datetime import datetime, timedelta
from collections import defaultdict

class ExportFormats:
    """Handle multiple export formats"""

    def __init__(self, parent, tracker):
        self.parent = parent
        self.tracker = tracker

    def create_export_ui(self, frame):
        """Create export UI"""
        # Clear frame
        for widget in frame.winfo_children():
            widget.destroy()

        # Header
        header = ctk.CTkLabel(
            frame,
            text="📤 Export Data",
            font=ctk.CTkFont(size=32, weight="bold")
        )
        header.pack(pady=20)

        # Main content
        content_frame = ctk.CTkFrame(frame)
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Export options
        options_frame = ctk.CTkFrame(content_frame)
        options_frame.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(
            options_frame,
            text="Select Date Range",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(anchor="w", padx=10, pady=(10, 5))

        # Date range selector
        range_frame = ctk.CTkFrame(options_frame, fg_color="transparent")
        range_frame.pack(fill="x", padx=10, pady=10)

        self.date_range_var = ctk.StringVar(value="last_7_days")
        date_options = [
            ("Last 7 Days", "last_7_days"),
            ("Last 30 Days", "last_30_days"),
            ("This Month", "this_month"),
            ("Last Month", "last_month"),
            ("All Time", "all_time")
        ]

        for text, value in date_options:
            ctk.CTkRadioButton(
                range_frame,
                text=text,
                variable=self.date_range_var,
                value=value,
                font=ctk.CTkFont(size=14)
            ).pack(anchor="w", pady=5, padx=20)

        # Format selection
        format_frame = ctk.CTkFrame(content_frame)
        format_frame.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(
            format_frame,
            text="Select Export Format",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(anchor="w", padx=10, pady=(10, 5))

        # Export format buttons
        btn_frame = ctk.CTkFrame(format_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=20)

        formats = [
            ("📄 CSV", self.export_csv, "#4CAF50"),
            ("📋 JSON", self.export_json, "#2196F3"),
            ("📊 Excel", self.export_excel, "#FF9800"),
            ("📑 PDF", self.export_pdf, "#F44336")
        ]

        for text, command, color in formats:
            ctk.CTkButton(
                btn_frame,
                text=text,
                command=command,
                height=60,
                width=150,
                font=ctk.CTkFont(size=16, weight="bold"),
                fg_color=color,
                hover_color=self.darken_color(color)
            ).pack(side="left", padx=10, expand=True)

        # Info
        info_frame = ctk.CTkFrame(content_frame)
        info_frame.pack(fill="x", padx=20, pady=10)

        info_text = """
        📄 CSV - Compatible with Excel, Google Sheets
        📋 JSON - Raw data, programming-friendly
        📊 Excel - Formatted spreadsheet with charts
        📑 PDF - Professional report format
        """

        ctk.CTkLabel(
            info_frame,
            text=info_text,
            font=ctk.CTkFont(size=12),
            text_color="gray",
            justify="left"
        ).pack(padx=20, pady=20)

    def get_date_range(self):
        """Get selected date range"""
        range_type = self.date_range_var.get()
        today = datetime.now()

        if range_type == "last_7_days":
            start_date = today - timedelta(days=7)
        elif range_type == "last_30_days":
            start_date = today - timedelta(days=30)
        elif range_type == "this_month":
            start_date = today.replace(day=1)
        elif range_type == "last_month":
            first_day_this_month = today.replace(day=1)
            start_date = (first_day_this_month - timedelta(days=1)).replace(day=1)
        else:  # all_time
            start_date = datetime(2000, 1, 1)

        return start_date

    def get_filtered_data(self):
        """Get data for selected date range"""
        start_date = self.get_date_range()
        filtered_data = {}

        for date_str, day_data in self.tracker.data.items():
            try:
                date = datetime.strptime(date_str, "%Y-%m-%d")
                if date >= start_date:
                    filtered_data[date_str] = day_data
            except:
                continue

        return filtered_data

    def export_csv(self):
        """Export to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"time_tracker_{datetime.now().strftime('%Y%m%d')}.csv"
            )

            if not filename:
                return

            data = self.get_filtered_data()

            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Date", "Category", "Hours", "Project"])

                for date_str, day_data in sorted(data.items()):
                    for category, hours in day_data.items():
                        if category not in ['date', 'session_duration', 'idle_time', 'projects']:
                            project = ""
                            if 'projects' in day_data:
                                for proj, proj_hours in day_data['projects'].items():
                                    if proj_hours > 0:
                                        project = proj
                                        break
                            writer.writerow([date_str, category, f"{hours:.2f}", project])

            messagebox.showinfo("Success", f"Data exported to CSV:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV:\n{str(e)}")

    def export_json(self):
        """Export to JSON"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile=f"time_tracker_{datetime.now().strftime('%Y%m%d')}.json"
            )

            if not filename:
                return

            data = self.get_filtered_data()

            with open(filename, 'w', encoding='utf-8') as f:
                json.dump({
                    "export_date": datetime.now().isoformat(),
                    "date_range": self.date_range_var.get(),
                    "data": data
                }, f, indent=2)

            messagebox.showinfo("Success", f"Data exported to JSON:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export JSON:\n{str(e)}")

    def export_excel(self):
        """Export to Excel"""
        try:
            # Check if openpyxl is available
            try:
                from openpyxl import Workbook
                from openpyxl.chart import PieChart, Reference
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                messagebox.showwarning(
                    "Missing Library",
                    "Excel export requires 'openpyxl' library.\n\n"
                    "Install it with: pip install openpyxl\n\n"
                    "Falling back to CSV export..."
                )
                self.export_csv()
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"time_tracker_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )

            if not filename:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Time Tracking Data"

            # Headers
            headers = ["Date", "Category", "Hours", "Project"]
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data
            data = self.get_filtered_data()
            for date_str, day_data in sorted(data.items()):
                for category, hours in day_data.items():
                    if category not in ['date', 'session_duration', 'idle_time', 'projects']:
                        project = ""
                        if 'projects' in day_data:
                            for proj, proj_hours in day_data['projects'].items():
                                if proj_hours > 0:
                                    project = proj
                                    break
                        ws.append([date_str, category, round(hours, 2), project])

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            summary = wb.create_sheet("Summary")
            summary.append(["Category", "Total Hours"])

            category_totals = defaultdict(float)
            for date_str, day_data in data.items():
                for category, hours in day_data.items():
                    if category not in ['date', 'session_duration', 'idle_time', 'projects']:
                        category_totals[category] += hours

            for category, total in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
                summary.append([category, round(total, 2)])

            # Style summary
            for cell in summary[1]:
                cell.fill = header_fill
                cell.font = header_font

            wb.save(filename)
            messagebox.showinfo("Success", f"Data exported to Excel:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel:\n{str(e)}")

    def export_pdf(self):
        """Export to PDF"""
        try:
            # Check if reportlab is available
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
            except ImportError:
                messagebox.showwarning(
                    "Missing Library",
                    "PDF export requires 'reportlab' library.\n\n"
                    "Install it with: pip install reportlab\n\n"
                    "Falling back to CSV export..."
                )
                self.export_csv()
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"time_tracker_{datetime.now().strftime('%Y%m%d')}.pdf"
            )

            if not filename:
                return

            doc = SimpleDocTemplate(filename, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor("#2196F3"),
                spaceAfter=30,
                alignment=TA_CENTER
            )

            story.append(Paragraph("⏱️ Time Tracker Report", title_style))
            story.append(Spacer(1, 0.2 * inch))

            # Date info
            info_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
            info_text += f"Date Range: {self.date_range_var.get().replace('_', ' ').title()}"
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 0.3 * inch))

            # Summary table
            data = self.get_filtered_data()
            category_totals = defaultdict(float)
            total_hours = 0

            for date_str, day_data in data.items():
                for category, hours in day_data.items():
                    if category not in ['date', 'session_duration', 'idle_time', 'projects']:
                        category_totals[category] += hours
                        total_hours += hours

            # Category summary
            story.append(Paragraph("Category Summary", styles['Heading2']))
            story.append(Spacer(1, 0.1 * inch))

            table_data = [["Category", "Hours", "Percentage"]]
            for category, hours in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
                percentage = (hours / total_hours * 100) if total_hours > 0 else 0
                table_data.append([category, f"{hours:.2f}h", f"{percentage:.1f}%"])

            table_data.append(["TOTAL", f"{total_hours:.2f}h", "100%"])

            table = Table(table_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2196F3")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#E3F2FD")),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(table)
            story.append(PageBreak())

            # Daily breakdown
            story.append(Paragraph("Daily Breakdown", styles['Heading2']))
            story.append(Spacer(1, 0.1 * inch))

            for date_str in sorted(data.keys(), reverse=True)[:30]:  # Last 30 days
                day_data = data[date_str]
                daily_total = sum(v for k, v in day_data.items()
                                if k not in ['date', 'session_duration', 'idle_time', 'projects'])

                if daily_total > 0:
                    story.append(Paragraph(f"<b>{date_str}</b> - {daily_total:.2f}h", styles['Normal']))

                    day_table = [["Category", "Hours"]]
                    for category, hours in sorted(day_data.items(), key=lambda x: x[1], reverse=True):
                        if category not in ['date', 'session_duration', 'idle_time', 'projects'] and hours > 0:
                            day_table.append([category, f"{hours:.2f}h"])

                    if len(day_table) > 1:
                        dt = Table(day_table, colWidths=[4*inch, 2*inch])
                        dt.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E3F2FD")),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                        ]))
                        story.append(dt)
                        story.append(Spacer(1, 0.2 * inch))

            doc.build(story)
            messagebox.showinfo("Success", f"Data exported to PDF:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF:\n{str(e)}")

    def darken_color(self, hex_color):
        """Darken a hex color"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        darker = tuple(max(0, int(c * 0.8)) for c in rgb)
        return f"#{darker[0]:02x}{darker[1]:02x}{darker[2]:02x}"
